# -*- coding:utf-8 -*-
# original author: songshanyuwu
# python             3.7.3



import re
from openpyxl import Workbook
import csv
import os



# # 获取当前文件目录
# def fileNameList():
#     # 获取脚本运行的目录
#     filePath = os.path.split(os.path.realpath(__file__))[0]
#     # print (filePath)

#     # 获取当前目录下的所有文件名
#     file_list = os.listdir(filePath)
#     # print(file_list)

#     # 生成文件绝对路径列表
#     fileList = []
#     for lm in file_list:
#         if '.py' in lm:
#             pass
#         else:
#             fileList.append(filePath + '\\' + lm)
#             # print(filePath + '\\' + lm)
#     return fileList


# 统计端口异常的IP、进程、次数
def tongji(fileName):
    resultdist = {}

    print(fileName)
    # with open(fileName,newline='',encoding='UTF-8') as csvfile:
    with open(fileName,newline='', encoding = 'GB2312') as csvfile:
        rows=csv.reader(csvfile)
        print(rows)

        for row in rows:
            print(','.join(row))   #row是列表类型
            result = re.findall(r"进程(?P<value>[A-Za-z0-9\.]+)", row[1])
            # 构建二重字典
            # 结构如下
            # {"10.118.249.161":{"DPMRA.exe":333,"LiveUpdate360.exe":2222}}

            # 构建进程计数字典
            tmpdist = {}
            for val in result:
                if val in tmpdist:
                    count = tmpdist[val] + 1
                    tmpdist.update({str(val):count})
                else:
                    tmpdist.update({str(val):1})
            # print(tmpdist)

            # 加入到结果字典
            if row[2] in resultdist:
                # 遍历字典去判断
                for key,value in tmpdist.items():
                    if key in resultdist[row[2]]:
                        count = value + resultdist[row[2]][key]
                        resultdist[row[2]].update({str(key):count})
                    else:
                        resultdist[row[2]].update({str(key):value})
            else:
                resultdist.update({row[2]:tmpdist})
    # 输出结果字典
    # print(resultdist)

    # 将字典输出到excel表格中
    # 要创建的xlsx名称
    dest_filename = r'C:\Users\songz\Desktop\新建文件夹\共享服务公司.xlsx'
    # 创建工作簿，设为当前，该标题名称
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "date"

    # 设置列名
    ws1.cell(row=1,column=1,value="IP地址")
    ws1.cell(row=1,column=2,value="进程名")
    ws1.cell(row=1,column=3,value="次数")

    # 循环数据写入内容
    i= 2
    for key,value in resultdist.items():
        ws1.cell(row=i, column=1, value=key)

        for k2,v2 in value.items():
            ws1.cell(row=i, column=2, value=k2)
            ws1.cell(row=i, column=3, value=v2)
            i += 1

    # 创建xlsx
    wb.save(filename=dest_filename)
    print('文件' + dest_filename + '生成完毕')



if __name__ == '__main__':
    # fileList = fileNameList()
    # print(fileList)
    fileName=r'C:\Users\songz\Desktop\新建文件夹\共享服务公司.csv'
    tongji(fileName)
    #for file in fileList:
    #    tongji(file)

