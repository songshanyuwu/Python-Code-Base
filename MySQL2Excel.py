# -*- coding:utf-8 -*-
# original author: 花有清香 
# https://blog.csdn.net/yang_ping_cai_niao/article/details/83896367
#
# python             3.7.3
# PyMySQL            0.9.3    链接数据库
# openpyxl           3.0.0    读写xlsx
# 在此我只贡献链接数据库和写入xlsx的代码

# 下面没有用上的，就注释掉了
#from fj.util import logger
#from openpyxl.compat import range
import pymysql.cursors
from openpyxl import Workbook
#from openpyxl.utils import get_column_letter

# 链接数据库的游标
connect = pymysql.Connect(
    host="localhost",
    port=3306,
    user='root',
    passwd='root',
    db='masscan',
    charset='utf8',
)
cursor = connect.cursor()

# 关闭数据库链接操作
def  clos_cursor():
    return cursor.close();


# 读取数据库数据
def query_all():
    select_sql = "select * from data"
    cursor.execute(select_sql);
    return cursor.fetchall();


# 关闭数据库链接操作
def  clos_cursor():
     cursor.close();
     connect.close()


def read_mysql_to_xlsx():
    #要创建的xlsx名称
    dest_filename = 'masscan_data.xlsx'
    #数字转IP地址的Python实现方法
    num2ip = lambda x: '.'.join([str(int((x/(256**i))%256)) for i in range(3,-1,-1)])
    #创建工作簿，设为当前，该标题名称
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "date"

    # 设置列名
    ws1.cell(row=1,column=1,value="id")
    ws1.cell(row=1,column=2,value="ip")
    ws1.cell(row=1,column=3,value="port_id")
    ws1.cell(row=1,column=4,value="scanned_ts")
    ws1.cell(row=1,column=5,value="protocol")
    ws1.cell(row=1,column=6,value="state")
    ws1.cell(row=1,column=7,value="reason")
    ws1.cell(row=1,column=8,value="reason_ttl")
    ws1.cell(row=1,column=9,value="service")
    ws1.cell(row=1,column=10,value="banner")
    ws1.cell(row=1,column=11,value="title")

    # 循环数据写入内容
    # 读取数据库的数据
    masscan_date_list = query_all()
    #                                     这里要+2，否则输出结果少最后一行哈
    for i in range(2,len(masscan_date_list)+2):
        #                                                这里i-2，否则输出少第一行┭┮﹏┭┮
        ws1.cell(row=i, column=1, value=masscan_date_list[i-2][0])
        ws1.cell(row=i, column=2, value=num2ip(masscan_date_list[i-2][1]))
        ws1.cell(row=i, column=3, value=masscan_date_list[i-2][2])
        ws1.cell(row=i, column=4, value=masscan_date_list[i-2][3])
        ws1.cell(row=i, column=5, value=masscan_date_list[i-2][4])
        ws1.cell(row=i, column=6, value=masscan_date_list[i-2][5])
        ws1.cell(row=i, column=7, value=masscan_date_list[i-2][6])
        ws1.cell(row=i, column=8, value=masscan_date_list[i-2][7])
        ws1.cell(row=i, column=9, value=masscan_date_list[i-2][8])
        ws1.cell(row=i, column=10, value=masscan_date_list[i-2][9])
        ws1.cell(row=i, column=11, value=masscan_date_list[i-2][10])

    # 创建xlsx
    wb.save(filename=dest_filename)



if __name__ == '__main__':
    read_mysql_to_xlsx()



