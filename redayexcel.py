# -*- coding:utf-8 -*-
#original author:injetlee 
# https://github.com/injetlee/Python
#
# python             3.7.3
# openpyxl           3.0.0
#

# 效果：生成一个Excel文件，其中第一个表是40*10的随机数方阵；第二个表是指定位置写入值；第三个表是列字母和数字之间的转换。
import random
from openpyxl import Workbook

# 坑一： compat在openpyxl 3.0中似乎没有了，我这里就注释掉了
#from openpyxl.compat import range

# 坑二： get_column_letter方法已经在openpyxl 的2.4版本中重写了，从cell移到了utils。要在openpyxl.utils 中导入才有效果
# from openpyxl.cell  import get_column_letter
from openpyxl.utils  import get_column_letter

wb = Workbook()
dest_filename = 'empty_book1.xlsx'
ws1 = wb.active  # 第一个表
ws1.title = "range names"  # 第一个表命名
# 遍历第一个表的1到40行，赋值一个600内的随机数
for row in range(1, 40):
    # 坑三：这里报类型错误，一定注意值必须是list、tuple、range或generator或dict.
    #ws1.append(range(60))
    col_value = []
    # 我这里只假设10列的随机数
    for i in range(10):
        col_value.append(random.randint(0,600)) 
    ws1.append(col_value)
ws2 = wb.create_sheet(title="Pi")
ws2['F5'] = 3.1415926
ws3 = wb.create_sheet(title="Data")
for row in range(5, 15):
    for col in range(5, 15):
        _ = ws3.cell(column=col, row=row, value="%s" % get_column_letter(col))
wb.save(filename=dest_filename)
